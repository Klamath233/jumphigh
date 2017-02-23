from openpyxl import load_workbook
import os
import json
from sklearn import linear_model
from sklearn.kernel_ridge import KernelRidge
from sklearn.metrics import mean_absolute_error
from sklearn.model_selection import cross_val_predict
from sklearn.ensemble import RandomForestRegressor
from sklearn.model_selection import LeaveOneOut
from sklearn import preprocessing
import matplotlib.pyplot as plt


def parse_all():
    """Returns a list of dictionaries representing the data parsed from /data directory."""
    file_names = map(lambda x: ("../data/" + x, x.split(".")[0]), os.listdir("../data"))
    workbooks = map(lambda x: (load_workbook(x[0]), x[1]),
                    file_names)
    sheets = map(lambda x: (x[0]["Sheet1"], x[1]),
                 workbooks)
    dataset = [{"id": sheet[1],
                "jump": sheet[0]["C2"].value,
                "age": sheet[0]["C4"].value,
                "exercise": sheet[0]["C5"].value,
                "competitive": sheet[0]["C6"].value,
                "height": sheet[0]["C7"].value,
                "weight": sheet[0]["C8"].value,
                "gender": sheet[0]["C9"].value,
                "injury": sheet[0]["C10"].value,
                "color": sheet[0]["C11"].value,
                "sleep": sheet[0]["C12"].value,
                "race": sheet[0]["C13"].value}
                for sheet in sheets]
    return dataset

def plot():
    xset = []
    yset = []
    mydict = parse_all()
    gender = []
    bf = []
    for i, sample in enumerate(mydict):
        tmpH = sample["height"] * 0.025
        tmpHSq = tmpH * tmpH
        BMi = (sample["weight"] * 0.45)  / tmpHSq
        BF = 1.2 *BMi + 0.23 * sample["age"] - 10.8 * (sample["gender"] - 1) - 5.4
        y = sample["jump"]
        if sample["gender"] == 2:
            gender.append(sample["competitive"])
            yset.append(y)
            bf.append(BF)
    plt.plot(gender,yset, "r*",bf,yset, "b*")
    plt.show()

def generate_test_samples():
    file_names = map(lambda x: ("../testdata/" + x, x.split(".")[0]), os.listdir("../testdata"))
    workbooks = map(lambda x: (load_workbook(x[0]), x[1]),
                    file_names)
    sheets = map(lambda x: (x[0]["Sheet1"], x[1]),
                 workbooks)
    mydict = [{"id": sheet[1],
                "age": sheet[0]["C4"].value,
                "exercise": sheet[0]["C5"].value,
                "competitive": sheet[0]["C6"].value,
                "height": sheet[0]["C7"].value,
                "weight": sheet[0]["C8"].value,
                "gender": sheet[0]["C9"].value,
                "injury": sheet[0]["C10"].value,
                "color": sheet[0]["C11"].value,
                "sleep": sheet[0]["C12"].value,
                "race": sheet[0]["C13"].value}
               for sheet in sheets]
    xset = []
    for sample in mydict:
        tmpH = sample["height"] * 0.025
        tmpHSq = tmpH * tmpH
        BMi = (sample["weight"] * 0.45)  / tmpHSq
        BF = 1.2 *BMi + 0.23 * sample["age"] - 10.8 * (sample["gender"] - 1) - 5.4
        x = [sample["gender"] -1, BF, sample["exercise"], sample["race"]]
        xset.append(x)
    return xset

def createmodel():
    xset = []
    yset = []
    mydict = parse_all()
    for sample in mydict:
        if sample["id"] == "T12" or sample["id"] == "T6" or sample["id"] == "T20" or sample["id"] == "T7":
            continue
        tmpH = sample["height"] * 0.025
        tmpHSq = tmpH * tmpH
        BMi = (sample["weight"] * 0.45)  / tmpHSq
        BF = 1.2 *BMi + 0.23 * sample["age"] - 10.8 * (sample["gender"] - 1) - 5.4

        x = [sample["gender"] -1, BF, sample["exercise"], sample["race"]]
        y = sample["jump"]
        xset.append(x)
        yset.append(y)

    xset = preprocessing.normalize(xset, norm='l2')
    reg = KernelRidge(kernel="poly", alpha=1, degree=1)
    reg = RandomForestRegressor(n_estimators=10)

    loo = LeaveOneOut()
    loo.get_n_splits(xset)
    predicted = []
    for train, test in loo.split(xset):
        trainset =  [xset[i] for i in train]
        labelset = [yset[i] for i in train]
        reg.fit(trainset,labelset)
        ts = test[0]
        predicted_y = reg.predict(xset[ts])
        predicted.append(predicted_y)
    predicted = cross_val_predict(reg, xset, yset, cv=22)
    print(mean_absolute_error(yset, predicted))
    reg.fit(xset,yset)
    testsamples = generate_test_samples()
    res = reg.predict[testsamples]
    print(res)


#plot()
createmodel()
